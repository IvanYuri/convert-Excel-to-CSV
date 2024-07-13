import os
from pyxlsb import open_workbook as open_xlsb_workbook
import openpyxl
import csv
from tkinter import Tk, Listbox, Button, MULTIPLE, END, Toplevel, Label
from tkinter.filedialog import askopenfilename, askdirectory

def select_sheets(sheetnames):
    selected_sheets = []

    def on_select():
        nonlocal selected_sheets
        selected_sheets = [sheetnames[i] for i in listbox.curselection()]
        top.destroy()

    top = Toplevel()
    top.title("Selecione as abas para converter")

    Label(top, text="Selecione as abas para converter").pack()

    listbox = Listbox(top, selectmode=MULTIPLE)
    listbox.pack()

    for sheet in sheetnames:
        listbox.insert(END, sheet)

    Button(top, text="OK", command=on_select).pack()

    top.wait_window()
    return selected_sheets

def sanitize_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))  # Retornar valor sem ponto decimal se for inteiro
    value_str = str(value)
    value_str = value_str.replace(';', ',')  # Substituir ponto e vírgula por vírgula
    return value_str

def convert_xlsb_to_csv(xlsb_file_path, output_dir, selected_sheets):
    base_filename = os.path.splitext(os.path.basename(xlsb_file_path))[0]

    # Abrir o arquivo XLSB
    with open_xlsb_workbook(xlsb_file_path) as wb:
        # Iterar sobre as folhas selecionadas no arquivo XLSB
        for sheetname in selected_sheets:
            with wb.get_sheet(sheetname) as sheet:
                csv_file_path = os.path.join(output_dir, f"{base_filename}_{sheetname}.csv")
                with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
                    writer = csv.writer(csv_file, delimiter=';')

                    # Iterar sobre as linhas e colunas do XLSB e copiar os dados
                    for row in sheet.rows():
                        writer.writerow([sanitize_value(col.v) for col in row])

    print(f'Arquivos convertidos e salvos no diretório {output_dir}')

def convert_xlsx_to_csv(xlsx_file_path, output_dir, selected_sheets):
    base_filename = os.path.splitext(os.path.basename(xlsx_file_path))[0]

    # Abrir o arquivo XLSX
    wb = openpyxl.load_workbook(xlsx_file_path, read_only=True)
    # Iterar sobre as folhas selecionadas no arquivo XLSX
    for sheetname in selected_sheets:
        sheet = wb[sheetname]
        csv_file_path = os.path.join(output_dir, f"{base_filename}_{sheetname}.csv")
        with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
            writer = csv.writer(csv_file, delimiter=';')

            # Iterar sobre as linhas e colunas do XLSX e copiar os dados
            for row in sheet.iter_rows(values_only=True):
                writer.writerow([sanitize_value(cell) for cell in row])

    print(f'Arquivos convertidos e salvos no diretório {output_dir}')

# Ocultar a janela principal do Tkinter
root = Tk()
root.withdraw()

# Caixa de diálogo para selecionar o arquivo XLSB ou XLSX
file_path = askopenfilename(
    title="Selecione o arquivo XLSB ou XLSX, seleciona na parte de baixo",
    filetypes=[("Excel Binary Workbook", ".xlsb"), ("Excel Workbook", ".xlsx")]
)

# Verificar se o arquivo foi selecionado
if file_path:
    # Verificar a extensão do arquivo e abrir para obter as folhas disponíveis
    if file_path.endswith('.xlsb'):
        with open_xlsb_workbook(file_path) as wb:
            sheetnames = wb.sheets
    elif file_path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheetnames = wb.sheetnames
    else:
        print("Formato de arquivo não suportado.")
        exit()

    # Caixa de diálogo para selecionar as folhas
    selected_sheets = select_sheets(sheetnames)

    if selected_sheets:
        # Caixa de diálogo para selecionar o diretório para salvar os arquivos CSV
        output_dir = askdirectory(
            title="Selecione o diretório para salvar os arquivos CSV"
        )

        # Verificar se o diretório foi selecionado
        if output_dir:
            # Chamar a função de conversão apropriada
            if file_path.endswith('.xlsb'):
                convert_xlsb_to_csv(file_path, output_dir, selected_sheets)
            elif file_path.endswith('.xlsx'):
                convert_xlsx_to_csv(file_path, output_dir, selected_sheets)
        else:
            print("Seleção de diretório para CSV cancelada.")
    else:
        print("Nenhuma folha aba para conversão.")
else:
    print("Seleção de arquivo cancelada.")
